from io import BytesIO
from django.core.mail import EmailMultiAlternatives
from django.contrib.sites.shortcuts import get_current_site
from django.http import HttpResponse
import xlwt
from datetime import datetime
from django.shortcuts import get_object_or_404
from django.template.loader import get_template, render_to_string
from xhtml2pdf import pisa
from adminpage.models import Notification, Izinler, ProductSellStatistic
from ecommerce.settings import EMAIL_HOST_USER
from user_accounts.models import User
from django.utils.html import strip_tags
from openpyxl import load_workbook

def exportExcel(filename, sheetname, columns, rows):
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = f'attachment; filename={filename}-' + str(datetime.now()) + '.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(f'{sheetname}')
    row_num = 0
    font_style = xlwt.XFStyle()
    font_style.font.bold = True
    columns = columns
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)
    font_style = xlwt.XFStyle()
    rows = rows
    for row in rows:
        row_num += 1

        for col_num in range(len(row)):
            ws.write(row_num, col_num, str(row[col_num]), font_style)
    wb.save(response)

    return response


def exportPdf(template_name, context_dict={}):
    template = get_template(template_name)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("utf-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def notReadNotification():
    notify = Notification.objects.filter(is_read=False).count()
    return notify


def readNotification():
    notify = Notification.objects.filter(is_read=False)
    return notify


def createNotification(type, title, detail):
    notify = Notification.objects.create(noti_type=type, title=title, detail=detail)
    notify.save()
    response = 'Create Notify'
    return response


def izinSor(user_id):
    user = User.objects.get(id=user_id)
    izin = Izinler.objects.filter(user=user).last()
    return izin


def sendAccountVerificationEmail(request, first_name, last_name, email, otp):
    try:

        my_subject = "Lütfen hesabınızı doğrulayınız."
        my_recipient = email

        current_site = get_current_site(request)

        html_message = render_to_string("backend/email/account_verification_email.html", {
            'first_name': first_name,
            'last_name': last_name,
            'domain': current_site,
            'otp': otp
        })

        plain_message = strip_tags(html_message)

        message = EmailMultiAlternatives(
            subject=my_subject,
            body=plain_message,
            from_email=EMAIL_HOST_USER,
            to=[my_recipient]
        )

        message.attach_alternative(html_message, "text/html")
        message.send()

        msg = 'success'
        return msg
    except:
        msg = 'failed'
        return msg


def sendOrderInfoEmail(platform, email, order):
    try:
        my_subject = "Yeni bir sipariş var!"
        my_recipient = email

        if platform == "Trendyol":
            html_message = render_to_string("backend/email/order.html", {
                'siparis_no': order.order_number,
                'quantity': order.quantity,
                'title': order.title,
                'unit_price': order.unit_price,
                'sales_amount': order.sales_amount,
                'order_date': order.order_date,
                'siparis_id': order.id,
            })
            plain_message = strip_tags(html_message)

            message = EmailMultiAlternatives(
                subject=my_subject,
                body=plain_message,
                from_email=EMAIL_HOST_USER,
                to=[my_recipient]
            )

            message.attach_alternative(html_message, "text/html")
            message.send()
        elif platform == "TredyShop":
            html_message = render_to_string("backend/email/order.html", {
                'siparis_no': order.order_number,
                'sales_amount': order.order_total,
                'order_date': order.created_at
            })
            plain_message = strip_tags(html_message)

            message = EmailMultiAlternatives(
                subject=my_subject,
                body=plain_message,
                from_email=EMAIL_HOST_USER,
                to=[my_recipient]
            )

            message.attach_alternative(html_message, "text/html")
            message.send()
        msg = 'success'
        return msg
    except:
        msg = 'failed'
        return msg


def customerTredyShopCreateOrder(request, email, address, order_list, order, total, grand_total):
    my_subject = "Yeni bir sipariş var!"
    my_recipient = email

    if order.delivery_price == 0:
        delivery_price = "Ücretsiz"
    else:
        delivery_price = order.delivery_price

    vade_farki = order.order_total - order.order_amount

    html_message = render_to_string("backend/email/customer_order.html", {
        'siparis_no': order.order_number,
        'order_list': order_list,
        'sales_amount': order.order_total,
        'order_date': order.created_at,
        'address': address,
        'order_list_count': len(order_list),
        'total': total,
        'grand_total': grand_total,
        'delivery_price': delivery_price,
        'vade_farki': vade_farki,
    })
    plain_message = strip_tags(html_message)

    message = EmailMultiAlternatives(
        subject=my_subject,
        body=plain_message,
        from_email=EMAIL_HOST_USER,
        to=[my_recipient]
    )

    message.attach_alternative(html_message, "text/html")
    message.send()

    msg = 'success'
    return msg


def customerTredyShopDeliveryOrder(request, email, address, order_list, order, total, grand_total):
    my_subject = "Siparişin Kargolandı"
    my_recipient = email

    if order.delivery_price == 0:
        delivery_price = "Ücretsiz"
    else:
        delivery_price = order.delivery_price

    vade_farki = order.order_total - order.order_amount

    html_message = render_to_string("backend/email/cargo_notification.html", {
        'siparis_no': order.order_number,
        'order_list': order_list,
        'sales_amount': order.order_total,
        'order_date': order.created_at,
        'address': address,
        'order_list_count': len(order_list),
        'total': total,
        'grand_total': grand_total,
        'delivery_price': delivery_price,
        'vade_farki': vade_farki,
    })
    plain_message = strip_tags(html_message)

    message = EmailMultiAlternatives(
        subject=my_subject,
        body=plain_message,
        from_email=EMAIL_HOST_USER,
        to=[my_recipient]
    )

    message.attach_alternative(html_message, "text/html")
    message.send()

    msg = 'success'
    return msg


def customerTredyShopExtraditionRequestOrder(request, email, order_number):
    my_subject = "İade talebiniz alındı."
    my_recipient = email

    html_message = render_to_string("backend/email/customer_extradition_request.html", {
        'siparis_no': order_number,
    })
    plain_message = strip_tags(html_message)

    message = EmailMultiAlternatives(
        subject=my_subject,
        body=plain_message,
        from_email=EMAIL_HOST_USER,
        to=[my_recipient]
    )

    message.attach_alternative(html_message, "text/html")
    message.send()

    msg = 'success'
    return msg


def sendExtraditionRequestInfoEmail(request, email, order_number, siparis_tarihi, iade_tarihi, urun_id):
    try:
        my_subject = "Yeni bir iade talebin var!"
        my_recipient = email

        html_message = render_to_string("backend/email/extradition_request.html", {
            'siparis_no': order_number,
            'order_date': siparis_tarihi,
            'iade_tarihi': iade_tarihi,
            'urun_id': urun_id,
        })
        plain_message = strip_tags(html_message)

        message = EmailMultiAlternatives(
            subject=my_subject,
            body=plain_message,
            from_email=EMAIL_HOST_USER,
            to=[my_recipient]
        )

        message.attach_alternative(html_message, "text/html")
        message.send()

        msg = 'success'
        return msg

    except:
        msg = 'failed'
        return msg


def productStatistic(barcode, title, quantity, satis, harcama=None):
    if ProductSellStatistic.objects.filter(barcode=barcode).count() > 0:
        statistic = get_object_or_404(ProductSellStatistic, barcode=barcode)
        statistic.sell_count += quantity
        statistic.satis += float(satis)
        statistic.save()
    else:
        ProductSellStatistic.objects.create(barcode=barcode, name=title, sell_count=quantity,
                                            satis=satis, maliyet=harcama)

def read_excel(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    data = list()
    for row in ws.iter_rows(min_row=2):
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        data.append(row_data)

    return data